"""
chatbot_evaluator.py
====================
Silent background evaluator for the First Aid Chatbot.

Logs every interaction to evaluation_results.xlsx in the project folder.
The file is created and updated automatically after every interaction.
No UI, no Streamlit dependency.

Usage (in streamlit_app.py)
---------------------------
    from chatbot_evaluator import record_interaction

    import time
    t0 = time.perf_counter()
    try:
        bot_response = chatbot.get_response(user_input)
        error = None
    except Exception as e:
        bot_response = ""
        error = str(e)
    elapsed = time.perf_counter() - t0

    record_interaction(
        user_input    = user_input,
        bot_response  = bot_response,
        response_time = elapsed,
        severity      = severity,       # 'critical' / 'urgent' / 'moderate' / 'normal'
        match_found   = match_found,    # bool
        detected_lang = detected_lang,  # e.g. 'en', 'ar', 'fr'
        input_lang    = input_lang,
        error         = error,
    )
"""

import datetime
import os
import tempfile

# ── Module-level state ────────────────────────────────────────────────────────
_session_start: datetime.datetime = datetime.datetime.now()
_interactions: list = []

# On Streamlit Cloud the app directory may be read-only; fall back to the
# system temp directory so the Excel file can always be written.
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_WRITABLE_DIR = _BASE_DIR if os.access(_BASE_DIR, os.W_OK) else tempfile.gettempdir()
_EXCEL_PATH = os.path.join(_WRITABLE_DIR, 'evaluation_results.xlsx')

print(f"[Evaluator] Initialised. Excel will be saved to: {_EXCEL_PATH}")


# ── Public API ────────────────────────────────────────────────────────────────

def record_interaction(
    user_input: str,
    bot_response: str,
    response_time: float,
    severity: str,
    match_found: bool,
    detected_lang: str = 'en',
    input_lang: str = 'en',
    error: str | None = None,
) -> None:
    """Record one user→bot interaction and immediately update the Excel file.

    Parameters
    ----------
    user_input    : raw text the user typed
    bot_response  : text the chatbot returned
    response_time : wall-clock seconds taken to generate the response
    severity      : one of 'critical', 'urgent', 'moderate', 'normal'
    match_found   : True if the chatbot found a knowledge-base match
    detected_lang : language code detected for this message
    input_lang    : expected language code (usually same as detected_lang)
    error         : exception message if one occurred, else None
    """
    _interactions.append({
        'timestamp':     datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'user_input':    user_input,
        'bot_response':  bot_response,
        'response_time': round(response_time, 3),
        'detected_lang': detected_lang,
        'input_lang':    input_lang,
        'lang_correct':  detected_lang == input_lang,
        'severity':      severity,
        'match_found':   match_found,
        'error':         error or '',
    })
    print(f"[Evaluator] Interaction #{len(_interactions)} recorded — saving to Excel...")
    _save_excel()


# ── Internal helpers ──────────────────────────────────────────────────────────

def _get_summary() -> dict:
    """Compute aggregated metrics from all recorded interactions."""
    n = len(_interactions)
    if n == 0:
        return {
            'total_questions':       0,
            'session_duration_s':    0,
            'avg_response_time_s':   0.0,
            'min_response_time_s':   0.0,
            'max_response_time_s':   0.0,
            'language_accuracy_pct': 0.0,
            'lang_errors':           0,
            'match_rate_pct':        0.0,
            'matches_found':         0,
            'matches_failed':        0,
            'severity_detected_pct': 0.0,
            'no_severity_count':     0,
            'error_count':           0,
            'error_rate_pct':        0.0,
        }

    times       = [r['response_time'] for r in _interactions]
    lang_ok     = sum(1 for r in _interactions if r['lang_correct'])
    matches     = sum(1 for r in _interactions if r['match_found'])
    no_severity = sum(1 for r in _interactions if r['severity'] == 'normal')
    errors      = sum(1 for r in _interactions if r['error'])
    duration_s  = (datetime.datetime.now() - _session_start).total_seconds()

    return {
        'total_questions':       n,
        'session_duration_s':    round(duration_s, 1),
        'avg_response_time_s':   round(sum(times) / n, 3),
        'min_response_time_s':   round(min(times), 3),
        'max_response_time_s':   round(max(times), 3),
        'language_accuracy_pct': round(lang_ok / n * 100, 1),
        'lang_errors':           n - lang_ok,
        'match_rate_pct':        round(matches / n * 100, 1),
        'matches_found':         matches,
        'matches_failed':        n - matches,
        'severity_detected_pct': round((n - no_severity) / n * 100, 1),
        'no_severity_count':     no_severity,
        'error_count':           errors,
        'error_rate_pct':        round(errors / n * 100, 1),
    }


def _save_excel() -> None:
    """Rebuild and overwrite evaluation_results.xlsx from the in-memory list."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        print(f"[Evaluator] ERROR: openpyxl not installed — cannot save Excel. "
              f"Run: pip install openpyxl\n  Detail: {e}")
        return

    try:
        summary = _get_summary()
        wb = openpyxl.Workbook()

        # ── Colour palette ────────────────────────────────────────────────────
        HEADER_FILL = PatternFill('solid', fgColor='1F4E79')   # dark navy
        HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
        ALT_FILL    = PatternFill('solid', fgColor='EBF3FB')   # pale blue
        LABEL_FILL  = PatternFill('solid', fgColor='2E75B6')   # mid blue
        LABEL_FONT  = Font(color='FFFFFF', bold=True)
        TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
        CENTRE      = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        thin        = Side(style='thin', color='BDD7EE')
        BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _hdr(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTRE
            c.border = BORDER

        def _dat(ws, row, col, value, alt=False):
            c = ws.cell(row=row, column=col, value=value)
            if alt:
                c.fill = ALT_FILL
            c.alignment = LEFT
            c.border = BORDER

        # ════════════════════════════════════════════════════════════════════
        # Sheet 1 – Interactions
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Interactions'

        ws1.merge_cells('A1:J1')
        t1 = ws1['A1']
        t1.value = 'First Aid Chatbot — Interaction Log'
        t1.font = TITLE_FONT
        t1.alignment = CENTRE

        headers = [
            'Timestamp', 'User Input', 'Bot Response',
            'Response Time (s)', 'Detected Lang', 'Input Lang',
            'Lang Correct?', 'Severity', 'Match Found?', 'Error',
        ]
        for col, h in enumerate(headers, 1):
            _hdr(ws1, 2, col, h)

        for i, rec in enumerate(_interactions):
            row = i + 3
            alt = (i % 2 == 1)
            _dat(ws1, row, 1,  rec['timestamp'],                        alt)
            _dat(ws1, row, 2,  rec['user_input'],                        alt)
            _dat(ws1, row, 3,  rec['bot_response'],                      alt)
            _dat(ws1, row, 4,  rec['response_time'],                     alt)
            _dat(ws1, row, 5,  rec['detected_lang'],                     alt)
            _dat(ws1, row, 6,  rec['input_lang'],                        alt)
            _dat(ws1, row, 7,  'Yes' if rec['lang_correct'] else 'No',   alt)
            _dat(ws1, row, 8,  rec['severity'].upper(),                  alt)
            _dat(ws1, row, 9,  'Yes' if rec['match_found'] else 'No',    alt)
            _dat(ws1, row, 10, rec['error'],                             alt)

        for col, w in enumerate([20, 40, 60, 18, 14, 12, 14, 12, 13, 30], 1):
            ws1.column_dimensions[get_column_letter(col)].width = w
        ws1.row_dimensions[1].height = 28
        ws1.row_dimensions[2].height = 22
        ws1.freeze_panes = 'A3'

        # ════════════════════════════════════════════════════════════════════
        # Sheet 2 – Summary
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Summary')

        ws2.merge_cells('A1:C1')
        t2 = ws2['A1']
        t2.value = 'First Aid Chatbot — Evaluation Summary'
        t2.font = TITLE_FONT
        t2.alignment = CENTRE

        def _section(ws, row, label):
            ws.merge_cells(f'A{row}:C{row}')
            c = ws.cell(row=row, column=1, value=label)
            c.fill = LABEL_FILL
            c.font = LABEL_FONT
            c.alignment = LEFT
            c.border = BORDER

        def _row(ws, row, label, value, note=''):
            for col, val in enumerate([label, value, note], 1):
                _dat(ws, row, col, val, alt=(row % 2 == 0))
            ws.cell(row=row, column=2).alignment = Alignment(
                horizontal='center', vertical='center'
            )

        _hdr(ws2, 2, 1, 'Metric')
        _hdr(ws2, 2, 2, 'Value')
        _hdr(ws2, 2, 3, 'Notes')

        r = 3
        _section(ws2, r, '📋  Session Overview');  r += 1
        _row(ws2, r, 'Total Questions Asked', summary['total_questions']);              r += 1
        _row(ws2, r, 'Session Duration',      f"{summary['session_duration_s']} s");   r += 1

        r += 1
        _section(ws2, r, '⏱  Response Time');  r += 1
        _row(ws2, r, 'Average Response Time', f"{summary['avg_response_time_s']} s");  r += 1
        _row(ws2, r, 'Minimum Response Time', f"{summary['min_response_time_s']} s");  r += 1
        _row(ws2, r, 'Maximum Response Time', f"{summary['max_response_time_s']} s");  r += 1

        r += 1
        _section(ws2, r, '🌐  Language Accuracy');  r += 1
        _row(ws2, r, 'Language Accuracy', f"{summary['language_accuracy_pct']} %",
             'Response language matched input language'); r += 1
        _row(ws2, r, 'Language Errors',   summary['lang_errors'],
             'Times response was in wrong language'); r += 1

        r += 1
        _section(ws2, r, '🎯  Match Confidence');  r += 1
        _row(ws2, r, 'Match Rate',         f"{summary['match_rate_pct']} %",
             'Queries that found a knowledge-base match'); r += 1
        _row(ws2, r, 'Successful Matches', summary['matches_found']);  r += 1
        _row(ws2, r, 'Failed / Fallback',  summary['matches_failed'],
             'Returned clarification or "not found"'); r += 1

        r += 1
        _section(ws2, r, '🚨  Severity Detection');  r += 1
        _row(ws2, r, 'Severity Detected', f"{summary['severity_detected_pct']} %",
             'Responses with critical/urgent/moderate label'); r += 1
        _row(ws2, r, 'No Severity Count', summary['no_severity_count'],
             'Responses with no severity label (normal)'); r += 1

        r += 1
        _section(ws2, r, '❌  Errors');  r += 1
        _row(ws2, r, 'Total Errors', summary['error_count']); r += 1
        _row(ws2, r, 'Error Rate',   f"{summary['error_rate_pct']} %"); r += 1

        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 45
        ws2.row_dimensions[1].height = 28
        ws2.row_dimensions[2].height = 22
        ws2.freeze_panes = 'A3'

        # ── Write to disk ─────────────────────────────────────────────────────
        wb.save(_EXCEL_PATH)
        n = len(_interactions)
        print(f"[Evaluator] Logged interaction #{n} — saved to {_EXCEL_PATH}")

    except Exception as e:
        print(f"[Evaluator] ERROR saving Excel file: {type(e).__name__}: {e}")
