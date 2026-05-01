import streamlit as st
import datetime
import time
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from streamlit_autorefresh import st_autorefresh
from multilingual_interactive_chatbot import MultilingualInteractiveFirstAidChatbot
import auth
from chatbot_evaluator import record_interaction, _interactions, _get_summary



# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="LifeAid Assistant",
    page_icon="⛑️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
/* Dark background on every layer Streamlit uses for the sidebar */
section[data-testid="stSidebar"],
section[data-testid="stSidebar"] > div:first-child,
section[data-testid="stSidebar"] > div:first-child > div,
[data-testid="stSidebarContent"] {
    background-color: #1a1a2e !important;
}
section[data-testid="stSidebar"] {
    border-right: 2px solid #e63946 !important;
    min-width: 240px !important;
}
/* Text — specific selectors only, no wildcard */
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] small {
    color: #f0f0f0 !important;
}
/* Buttons */
section[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.22) !important;
    color: #f0f0f0 !important;
    width: 100% !important;
    text-align: left !important;
    border-radius: 6px !important;
    font-size: 13px !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.18) !important;
    color: #ffffff !important;
}
/* Horizontal rules */
section[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.2) !important;
}
</style>
""", unsafe_allow_html=True)
# ─── Auto-refresh ─────────────────────────────────────────────────────────────
if 'skip_refresh' not in st.session_state:
    st.session_state.skip_refresh = False

if st.session_state.get('active_timer'):
    if not st.session_state.get('timer_paused', False):
        if not st.session_state.skip_refresh:
            st_autorefresh(interval=1000, key="timer_refresh")
    else:
        st_autorefresh(interval=30000, key="timer_refresh")
    st.session_state.skip_refresh = False

# ─── Session state ────────────────────────────────────────────────────────────
for _k, _v in {
    'page': 'landing', 'logged_in': False, 'username': None,
    'chat_history': [], 'active_timer': None, 'timer_start_time': None,
    'timer_paused': False, 'timer_paused_remaining': None,
    'completion_sound_played': False, 'current_chat_saved': False,
    'dark_mode': True, 'show_settings': False, 'feedback_list': [],
    'show_emergency_numbers': False, 'feedback_key': 0, 'feedback_submitted': False,
    'show_eval_download': False,
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

if 'chatbot' not in st.session_state:
    with st.spinner('Initializing LifeAid...'):
        st.session_state.chatbot = MultilingualInteractiveFirstAidChatbot()

# ─── Theme ────────────────────────────────────────────────────────────────────
_DK = st.session_state.dark_mode
_is_chat = st.session_state.page == 'chat'

if _DK:
    _C = dict(
        bg="#0D1117", surface="#161B22", card="#21262D", border="#30363D",
        text="#E6EDF3", muted="#8B949E", primary="#58A6FF", sidebar="#161B22",
        input_bg="#0D1117", danger="#F85149", warning="#FFA657",
        success="#56D364", chat_user_bg="#1F6FEB",
        chat_bot_bg="#21262D", hero_card="rgba(255,255,255,0.07)",
        hero_border="rgba(255,255,255,0.12)",
    )
else:
    _C = dict(
        bg="#F2EFE9", surface="#EBE7E0", card="#FAF8F5", border="#D6D0C8",
        text="#1F2328", muted="#636C76", primary="#0969DA", sidebar="#EBE7E0",
        input_bg="#FAF8F5", danger="#CF222E", warning="#BC4C00",
        success="#1A7F37", chat_user_bg="#0969DA",
        chat_bot_bg="#EBE7E0", hero_card="rgba(255,255,255,0.15)",
        hero_border="rgba(255,255,255,0.25)",
    )


# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800;1,400&display=swap');

*, *::before, *::after {{
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
    box-sizing: border-box !important;
}}

#MainMenu, footer, header {{ visibility: hidden !important; }}
[data-testid="InputInstructions"] {{ display: none !important; }}

/* ── App shell ── */
.stApp {{
    background: {_C['bg']} !important;
    transition: background 0.25s ease !important;
}}
.main .block-container {{
    padding: 1.75rem 2.5rem 2rem 2.5rem !important;
    max-width: 100% !important;
}}

/* ── Typography ── */
/* No !important on broad selectors — would override intentional inline colors
   (e.g. white text inside dark-gradient hero cards). Scoped rules below
   handle the specific Streamlit containers that need forced colour. */
/* Headings rendered by Streamlit's markdown get the theme colour;
   inline-styled headings inside st.markdown HTML keep their own colour. */
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3,
[data-testid="stMarkdownContainer"] h4,
[data-testid="stMarkdownContainer"] h5,
[data-testid="stMarkdownContainer"] h6 {{
    color: {_C['text']} !important;
}}
h1, h2, h3, h4, h5, h6 {{
    letter-spacing: -0.025em !important;
    line-height: 1.25 !important;
}}
/* Streamlit's own markdown/text containers */
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] li,
[data-testid="stText"] p,
.stAlert p, .stAlert span,
[data-testid="stExpander"] p,
[data-testid="stExpander"] span,
[data-testid="stForm"] label,
[data-testid="stForm"] p {{
    color: {_C['text']} !important;
}}
/* Labels for native widgets (selectbox, text_input, etc.) */
label {{ color: {_C['text']} !important; }}

/* ── Buttons (base) ── */
.stButton > button {{
    background: {_C['card']} !important;
    color: {_C['text']} !important;
    border: 1px solid {_C['border']} !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    padding: 0.45rem 0.85rem !important;
    transition: all 0.18s cubic-bezier(.4,0,.2,1) !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    gap: 0.3rem !important;
    white-space: nowrap !important;
    cursor: pointer !important;
}}
.stButton > button:hover {{
    background: {_C['primary']}18 !important;
    border-color: {_C['primary']} !important;
    color: {_C['primary']} !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 14px {_C['primary']}28 !important;
}}
.stButton > button:active {{
    transform: translateY(0) !important;
    box-shadow: none !important;
}}
.stButton > button p {{ margin: 0 !important; color: inherit !important; }}

/* ── Text inputs ── */
[data-testid="stTextInput"] input {{
    background: {_C['input_bg']} !important;
    color: {_C['text']} !important;
    border: 1px solid {_C['border']} !important;
    border-radius: 8px !important;
    font-size: 14px !important;
    transition: border-color 0.15s, box-shadow 0.15s !important;
}}
[data-testid="stTextInput"] input:focus {{
    border-color: {_C['primary']} !important;
    box-shadow: 0 0 0 3px {_C['primary']}22 !important;
    outline: none !important;
}}
[data-testid="stTextInput"] label {{
    font-weight: 500 !important;
    font-size: 13px !important;
    color: {_C['muted']} !important;
    margin-bottom: 0.2rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
}}

/* ── Chat input — default Streamlit appearance, only text colour overridden ── */
[data-testid="stChatInput"] textarea,
[data-testid="stChatInput"] textarea:focus {{
    color: white !important;
    -webkit-text-fill-color: white !important;
}}
[data-testid="stChatInput"] textarea::placeholder {{
    color: rgba(255,255,255,0.55) !important;
    opacity: 1 !important;
}}

/* ── Chat messages ── */
[data-testid="stChatMessage"] {{
    animation: msgIn 0.22s cubic-bezier(.4,0,.2,1) !important;
    background: transparent !important;
    border: none !important;
    padding: 0.15rem 0 !important;
}}
@keyframes msgIn {{
    from {{ opacity: 0 !important; transform: translateY(10px) !important; }}
    to   {{ opacity: 1 !important; transform: translateY(0) !important; }}
}}

/* ── Progress bar ── */
[data-testid="stProgressBar"] {{
    border-radius: 99px !important;
    overflow: hidden !important;
    height: 8px !important;
    background: {_C['border']} !important;
}}
[data-testid="stProgressBar"] > div > div {{
    background: linear-gradient(90deg, {_C['primary']}, #7C3AED) !important;
    border-radius: 99px !important;
    transition: width 0.95s linear !important;
}}

/* ── Expander ── */
details[data-testid="stExpander"] {{
    border: 1px solid {_C['border']} !important;
    border-radius: 10px !important;
    background: {_C['card']} !important;
    overflow: hidden !important;
    margin-bottom: 0.5rem !important;
}}
details[data-testid="stExpander"] summary {{
    font-weight: 600 !important;
    font-size: 13px !important;
    color: {_C['text']} !important;
    padding: 0.65rem 1rem !important;
}}
details[data-testid="stExpander"] summary:hover {{
    background: {_C['primary']}10 !important;
}}

/* ── Form ── */
[data-testid="stForm"] {{
    border: 1px solid {_C['border']} !important;
    border-radius: 14px !important;
    background: {_C['card']} !important;
    padding: 2rem !important;
    box-shadow: 0 4px 24px rgba(0,0,0,0.07) !important;
}}

/* ── Dividers ── */
hr {{
    border: none !important;
    border-top: 1px solid {_C['border']} !important;
    margin: 0.75rem 0 !important;
}}

/* ── Alerts ── */
[data-testid="stAlert"] {{ border-radius: 9px !important; }}

/* ── Spinner ── */
[data-testid="stSpinner"] p {{ color: {_C['muted']} !important; font-size: 13px !important; }}

/* ── RTL / Arabic ── */
[lang="ar"], [dir="rtl"], .arabic-text {{
    direction: rtl !important;
    text-align: right !important;
    unicode-bidi: embed !important;
}}

/* ── Scrollbar ── */
::-webkit-scrollbar {{ width: 5px !important; height: 5px !important; }}
::-webkit-scrollbar-track {{ background: transparent !important; }}
::-webkit-scrollbar-thumb {{ background: {_C['border']} !important; border-radius: 99px !important; }}
::-webkit-scrollbar-thumb:hover {{ background: {_C['muted']} !important; }}

/* ── Animations ── */
@keyframes fadeUp {{
    from {{ opacity: 0 !important; transform: translateY(18px) !important; }}
    to   {{ opacity: 1 !important; transform: translateY(0) !important; }}
}}
@keyframes fadeIn {{
    from {{ opacity: 0 !important; }}
    to   {{ opacity: 1 !important; }}
}}
@keyframes pulse {{
    0%, 100% {{ opacity: 1 !important; }}
    50%       {{ opacity: 0.65 !important; }}
}}
@keyframes timerGlow {{
    0%, 100% {{ box-shadow: 0 0 0 0 {_C['primary']}40 !important; }}
    50%       {{ box-shadow: 0 0 0 8px {_C['primary']}00 !important; }}
}}

/* ── Hero cards — force white text over any scoped overrides ── */
.lifeaid-hero p,
.lifeaid-hero span,
.lifeaid-hero h1,
.lifeaid-hero h2,
.lifeaid-hero h3,
.lifeaid-hero div {{
    color: #F8FAFC !important;
}}

</style>
""", unsafe_allow_html=True)


# ─── Helper: section label ────────────────────────────────────────────────────
def _sb_label(text):
    st.markdown(f"""
    <p style="font-size:10px; font-weight:700; letter-spacing:0.1em;
       color:{_C['muted']}; text-transform:uppercase; margin:1.1rem 0 0.4rem 0;">
       {text}
    </p>""", unsafe_allow_html=True)


# ─── Logic helpers (unchanged) ────────────────────────────────────────────────
def clean_response(text):
    timer_instruction = ""
    if "⏱️ **Timer Available:**" in text:
        parts = text.split("⏱️ **Timer Available:**", 1)
        text = parts[0]
        timer_instruction = "⏱️ **Timer Available:**" + parts[1]
    if "Response Options:" in text:
        text = text.split("Response Options:")[0].strip()
    text = text.strip('"').strip("'")
    if timer_instruction:
        if not text.rstrip().endswith("━━━━━━━━━━━━━━━━━━━━━━"):
            text = text.rstrip() + "\n\n━━━━━━━━━━━━━━━━━━━━━━"
        text = text + "\n\n" + timer_instruction
    return text


def _build_eval_excel() -> bytes:
    """Build the evaluation workbook in memory and return raw bytes for download."""

    summary = _get_summary()
    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
    ALT_FILL    = PatternFill('solid', fgColor='EBF3FB')
    LABEL_FILL  = PatternFill('solid', fgColor='2E75B6')
    LABEL_FONT  = Font(color='FFFFFF', bold=True)
    TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
    CENTRE      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='BDD7EE')
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTRE;  c.border = BORDER

    def _dat(ws, row, col, value, alt=False):
        c = ws.cell(row=row, column=col, value=value)
        if alt:
            c.fill = ALT_FILL
        c.alignment = LEFT; c.border = BORDER

    def _section(ws, row, label):
        ws.merge_cells(f'A{row}:C{row}')
        c = ws.cell(row=row, column=1, value=label)
        c.fill = LABEL_FILL; c.font = LABEL_FONT
        c.alignment = LEFT; c.border = BORDER

    def _row(ws, row, label, value, note=''):
        for col, val in enumerate([label, value, note], 1):
            _dat(ws, row, col, val, alt=(row % 2 == 0))
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal='center', vertical='center')

    # ── Sheet 1: Interactions ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Interactions'
    ws1.merge_cells('A1:J1')
    t1 = ws1['A1']
    t1.value = 'First Aid Chatbot — Interaction Log'
    t1.font = TITLE_FONT; t1.alignment = CENTRE

    headers = [
        'Timestamp', 'User Input', 'Bot Response',
        'Response Time (s)', 'Detected Lang', 'Input Lang',
        'Lang Correct?', 'Severity', 'Match Found?', 'Error',
    ]
    for col, h in enumerate(headers, 1):
        _hdr(ws1, 2, col, h)

    for i, rec in enumerate(_interactions):
        row = i + 3
        alt = (i % 2 == 1)
        _dat(ws1, row, 1,  rec['timestamp'],                       alt)
        _dat(ws1, row, 2,  rec['user_input'],                       alt)
        _dat(ws1, row, 3,  rec['bot_response'],                     alt)
        _dat(ws1, row, 4,  rec['response_time'],                    alt)
        _dat(ws1, row, 5,  rec['detected_lang'],                    alt)
        _dat(ws1, row, 6,  rec['input_lang'],                       alt)
        _dat(ws1, row, 7,  'Yes' if rec['lang_correct'] else 'No',  alt)
        _dat(ws1, row, 8,  rec['severity'].upper(),                 alt)
        _dat(ws1, row, 9,  'Yes' if rec['match_found'] else 'No',   alt)
        _dat(ws1, row, 10, rec['error'],                            alt)

    for col, w in enumerate([20, 40, 60, 18, 14, 12, 14, 12, 13, 30], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[1].height = 28
    ws1.row_dimensions[2].height = 22
    ws1.freeze_panes = 'A3'

    # ── Sheet 2: Summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.merge_cells('A1:C1')
    t2 = ws2['A1']
    t2.value = 'First Aid Chatbot — Evaluation Summary'
    t2.font = TITLE_FONT; t2.alignment = CENTRE

    _hdr(ws2, 2, 1, 'Metric')
    _hdr(ws2, 2, 2, 'Value')
    _hdr(ws2, 2, 3, 'Notes')

    r = 3
    _section(ws2, r, 'Session Overview');  r += 1
    _row(ws2, r, 'Total Questions Asked', summary['total_questions']); r += 1
    _row(ws2, r, 'Session Duration', f"{summary['session_duration_s']} s"); r += 1

    r += 1
    _section(ws2, r, 'Response Time'); r += 1
    _row(ws2, r, 'Average Response Time', f"{summary['avg_response_time_s']} s"); r += 1
    _row(ws2, r, 'Minimum Response Time', f"{summary['min_response_time_s']} s"); r += 1
    _row(ws2, r, 'Maximum Response Time', f"{summary['max_response_time_s']} s"); r += 1

    r += 1
    _section(ws2, r, 'Language Accuracy'); r += 1
    _row(ws2, r, 'Language Accuracy', f"{summary['language_accuracy_pct']} %",
         'Response language matched input language'); r += 1
    _row(ws2, r, 'Language Errors', summary['lang_errors'],
         'Times response was in wrong language'); r += 1

    r += 1
    _section(ws2, r, 'Match Confidence'); r += 1
    _row(ws2, r, 'Match Rate', f"{summary['match_rate_pct']} %",
         'Queries that found a knowledge-base match'); r += 1
    _row(ws2, r, 'Successful Matches', summary['matches_found']); r += 1
    _row(ws2, r, 'Failed / Fallback', summary['matches_failed'],
         'Returned clarification or not found'); r += 1

    r += 1
    _section(ws2, r, 'Severity Detection'); r += 1
    _row(ws2, r, 'Severity Detected', f"{summary['severity_detected_pct']} %",
         'Responses with critical/urgent/moderate label'); r += 1
    _row(ws2, r, 'No Severity Count', summary['no_severity_count'],
         'Responses with no severity label (normal)'); r += 1

    r += 1
    _section(ws2, r, 'Errors'); r += 1
    _row(ws2, r, 'Total Errors', summary['error_count']); r += 1
    _row(ws2, r, 'Error Rate', f"{summary['error_rate_pct']} %"); r += 1

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 45
    ws2.row_dimensions[1].height = 28
    ws2.row_dimensions[2].height = 22
    ws2.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def detect_timer_need(response_text, user_input=''):
    import re
    response_lower = response_text.lower()
    user_lower = user_input.lower()

    # Bleeding questions must NEVER trigger any timer.
    _bleed_words = ['bleed', 'bleeding', 'blood loss', 'hemorrhage', 'haemorrhage']
    _is_bleeding = any(w in user_lower for w in _bleed_words)
    # Only suppress if the question isn't also clearly about CPR, burns, or choking.
    _non_bleed_emergency = any(w in user_lower for w in [
        'cpr', 'cardiac', 'heart attack', 'burn', 'chok', 'heimlich', 'thrust',
    ])
    if _is_bleeding and not _non_bleed_emergency:
        return None

    timers = {
        'cpr_cycle': {
            'keywords': ['cpr', 'chest compression', 'compressions'],
            'duration': 120, 'title': '❤️ CPR Cycle Timer',
            'instructions': '30 compressions, then 2 breaths - repeat',
        },
        'burn_cooling': {
            'keywords': ['cool', 'water', 'rinse'],
            'duration': 1200, 'title': '🔥 Burn Cooling Timer',
            'instructions': 'Keep burn under cool running water',
        },
        'choking_response': {
            'keywords': ['thrust', 'heimlich', 'back blows', 'wrap your arms'],
            'duration': 300, 'title': '🫁 Choking Response Timer',
            'instructions': 'Continue attempts until object dislodges',
        },
    }
    for timer_type, info in timers.items():
        if timer_type == 'burn_cooling':
            # Use word-boundary matching so "burning sensation" in a wound/bleeding
            # response does not falsely satisfy the burn guard.
            burn_in_response = bool(re.search(r'\bburn(s|ed)?\b', response_lower))
            burn_in_user = bool(re.search(r'\bburn(s|ed)?\b', user_lower))
            if (burn_in_response or burn_in_user) and any(k in response_lower for k in info['keywords']):
                return info
        else:
            if any(k in response_lower for k in info['keywords']):
                return info
    return None


def get_timer_instruction_translated(timer_title, language='en'):
    scroll_button = (
        f"""<a href="#timer-section" style="background:{_C['primary']};color:white;"""
        """padding:4px 10px;border-radius:12px;text-decoration:none;font-size:12px;"""
        """display:inline-block;margin-left:6px;">↑ Go to Timer</a>"""
    )
    t = {
        'en': f"⏱️ **Timer Available:** A {timer_title} is ready above. {scroll_button} Click '▶️ Resume' to start.",
        'ar': f"⏱️ **مؤقت متاح:** {timer_title} جاهز أعلاه. {scroll_button} انقر على '▶️ استئناف'.",
        'es': f"⏱️ **Temporizador disponible:** {timer_title} listo. {scroll_button} Haz clic en '▶️ Reanudar'.",
        'fr': f"⏱️ **Minuteur disponible:** {timer_title} prêt. {scroll_button} Cliquez sur '▶️ Reprendre'.",
        'de': f"⏱️ **Timer verfügbar:** {timer_title} bereit. {scroll_button} Klicken Sie auf '▶️ Fortsetzen'.",
        'pt': f"⏱️ **Temporizador disponível:** {timer_title} pronto. {scroll_button} Clique em '▶️ Retomar'.",
        'ru': f"⏱️ **Таймер доступен:** {timer_title} готов. {scroll_button} Нажмите '▶️ Продолжить'.",
        'ja': f"⏱️ **タイマー利用可能：** {timer_title} 準備完了。{scroll_button} 「▶️ 再開」をクリック。",
        'ko': f"⏱️ **타이머 사용 가능:** {timer_title} 준비됨。{scroll_button} '▶️ 재개' 클릭。",
        'it': f"⏱️ **Timer disponibile:** {timer_title} pronto. {scroll_button} Fare clic su '▶️ Riprendi'.",
        'nl': f"⏱️ **Timer beschikbaar:** {timer_title} klaar. {scroll_button} Klik op '▶️ Hervatten'.",
        'pl': f"⏱️ **Timer dostępny:** {timer_title} gotowy. {scroll_button} Kliknij '▶️ Wznów'.",
        'tr': f"⏱️ **Zamanlayıcı mevcut:** {timer_title} hazır. {scroll_button} '▶️ Devam Et' tıklayın.",
        'hi': f"⏱️ **टाइमर उपलब्ध:** {timer_title} तैयार। {scroll_button} '▶️ फिर से शुरू' क्लिक करें।",
        'id': f"⏱️ **Timer tersedia:** {timer_title} siap. {scroll_button} Klik '▶️ Lanjutkan'.",
        'vi': f"⏱️ **Bộ đếm thời gian:** {timer_title} sẵn sàng. {scroll_button} Nhấp '▶️ Tiếp tục'.",
        'th': f"⏱️ **ตัวจับเวลา:** {timer_title} พร้อม. {scroll_button} คลิก '▶️ ดำเนินการต่อ'.",
    }
    return t.get(language, t['en'])


def text_to_speech_button(text, message_id):
    from gtts import gTTS
    import os, tempfile, re

    clean_text = text.replace('*', '').replace('#', '').replace('_', '')
    clean_text = clean_text.replace('🚨', '').replace('⚠️', '').replace('⚠', '')
    clean_text = clean_text.replace('🟢', '').replace('🔵', '')
    clean_text = clean_text.replace('━━━━━━━━━━━━━━━━━━━━━━', '').replace('\n', ' ')
    clean_text = clean_text.replace('📋', '').replace('⏱️', '').replace('❤️', '') \
                            .replace('🔥', '').replace('🫁', '').replace('🫀', '')
    clean_text = re.sub(r'<[^>]+>', '', clean_text)
    clean_text = re.sub(r'\b[A-Z]{2,}\b', lambda m: m.group().title(), clean_text)
    clean_text = ' '.join(clean_text.split()).strip()

    lang = 'en'
    if any(c in text for c in ['ا', 'ب', 'ت', 'ث', 'ج', 'ح', 'خ', 'د', 'ذ', 'ر', 'ز', 'س', 'ش']):
        lang = 'ar'

    audio_key = f"audio_cache_{message_id}"
    play_clicked = st.button("🔊", key=f"tts_btn_{message_id}", help="Listen to this message")

    if play_clicked:
        if audio_key in st.session_state and st.session_state[audio_key]:
            st.audio(st.session_state[audio_key], format='audio/mp3', autoplay=True)
        else:
            with st.spinner('🔊 Generating audio...'):
                try:
                    temp_dir = tempfile.gettempdir()
                    audio_file = os.path.join(temp_dir, f"tts_{message_id}.mp3")
                    tts = gTTS(text=', ' + clean_text, lang=lang, slow=False)
                    tts.save(audio_file)
                    with open(audio_file, 'rb') as f:
                        audio_bytes = f.read()
                    st.session_state[audio_key] = audio_bytes
                    if os.path.exists(audio_file):
                        os.remove(audio_file)
                    st.audio(audio_bytes, format='audio/mp3', autoplay=True)
                except Exception as e:
                    st.error(f"Audio error: {e}")


# ─── Quick help handler ───────────────────────────────────────────────────────
def handle_quick_help(topic):
    _topics = {
        'bleeding': dict(
            query='What are the steps to control bleeding from a wound?',
            timer=None, default_sev='urgent',
        ),
        'choking': dict(
            query='How do I help a choking person?',
            timer=dict(duration=300, title='🫁 Choking Response Timer',
                       instructions='Continue attempts until object dislodges'),
            default_sev='critical',
        ),
        'cpr': dict(
            query='What should you do if someone has a cardiac arrest?',
            timer=dict(duration=120, title='❤️ CPR Cycle Timer',
                       instructions='30 compressions, then 2 breaths - repeat'),
            default_sev='critical',
        ),
        'burns': dict(
            query='Someone has a burn',
            timer=dict(duration=1200, title='🔥 Burn Cooling Timer',
                       instructions='Keep burn under cool running water'),
            default_sev='moderate',
        ),
    }
    cfg = _topics[topic]
    query = cfg['query']
    timer = cfg['timer']

    st.session_state.chat_history.append({'role': 'user', 'content': query})
    with st.spinner('Getting help...'):
        response = st.session_state.chatbot.get_response(query)

    severity = ('critical' if 'CRITICAL' in response
                else ('urgent' if 'URGENT' in response
                      else cfg['default_sev']))

    if timer:
        st.session_state.active_timer = timer
        st.session_state.timer_start_time = datetime.datetime.now()
        st.session_state.timer_paused = True
        st.session_state.timer_paused_remaining = timer['duration']
        user_lang = getattr(st.session_state.chatbot, 'user_language', 'en')
        response += "\n" + get_timer_instruction_translated(timer['title'], user_lang)

    st.session_state.chat_history.append(
        {'role': 'assistant', 'content': response, 'severity': severity})

    if st.session_state.logged_in:
        if not st.session_state.current_chat_saved:
            auth.save_user_chat(st.session_state.username, st.session_state.chat_history)
            st.session_state.current_chat_saved = True
        else:
            auth.update_user_chat(st.session_state.username, st.session_state.chat_history)
    st.rerun()



# ─── Landing page ─────────────────────────────────────────────────────────────
def show_landing_page():
    HB = "linear-gradient(135deg, #0F1B35 0%, #1a1060 40%, #6B21A8 100%)"
    HT = "#F8FAFC"
    HM = "rgba(248,250,252,0.72)"

    # ── Navbar ──
    st.markdown(
        f'<div style="text-align:center;padding:0.6rem 0 0.4rem;">'
        f'<p style="margin:0;font-size:26px;font-weight:700;'
        f'color:{_C["text"]};letter-spacing:-0.02em;">⛑️ LifeAid</p>'
        f'<p style="margin:0.2rem 0 0;font-size:16px;'
        f'color:{_C["muted"]};">Emergency First Aid · Multilingual AI</p>'
        f'</div>',
        unsafe_allow_html=True,
    )
    st.markdown(f'<hr style="border-color:{_C["border"]};margin:0.6rem 0 1.5rem;">', unsafe_allow_html=True)

    # ── Hero – two columns ──
    hero_l, hero_r = st.columns([3, 2], gap="large")

    with hero_l:
        # Left pane: gradient card with headline + tagline
        st.markdown(f"""
<div class="lifeaid-hero" style="background:{HB};border-radius:18px;padding:3.5rem 2.75rem 2.5rem;position:relative;overflow:hidden;">
  <div style="position:absolute;inset:0;opacity:0.06;background-image:radial-gradient(circle,white 1px,transparent 1px);background-size:26px 26px;"></div>
  <div style="position:relative;z-index:1;">
    <div style="display:inline-block;background:rgba(255,255,255,0.12);border:1px solid rgba(255,255,255,0.22);border-radius:99px;padding:0.3rem 0.9rem;font-size:11px;font-weight:600;color:{HM};letter-spacing:0.07em;margin-bottom:1.4rem;">
      ✦ AI-POWERED FIRST AID GUIDANCE
    </div>
    <h1 style="margin:0 0 1rem;font-size:52px;font-weight:800;color:{HT};line-height:1.06;letter-spacing:-0.035em;">
      Fast First Aid,<br>
      <span style="background:linear-gradient(90deg,#60A5FA,#C084FC);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;">Any Language.</span>
    </h1>
    <p style="font-size:16px;color:{HM};margin:0 0 1.75rem;line-height:1.7;font-weight:400;max-width:480px;">
      Clear, step-by-step emergency guidance when every second counts.
      Works in 15+ languages with built-in timers for CPR, burns, and choking.
    </p>
    <p style="margin:0;font-size:12px;color:{HM};opacity:0.8;">No account needed &nbsp;&middot;&nbsp; Free to use</p>
  </div>
</div>
""", unsafe_allow_html=True)

        # CTA buttons below the gradient card
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        b1, b2, _sp = st.columns([1.4, 1.4, 3])
        with b1:
            if st.button("Open Chat →", key="hero_open_chat", use_container_width=True):
                st.session_state.page = 'chat'
                st.rerun()
        with b2:
            if st.session_state.logged_in:
                if st.button("Logout", key="hero_logout", use_container_width=True):
                    st.session_state.logged_in = False
                    st.session_state.username = None
                    st.session_state.chat_history = []
                    st.rerun()
            else:
                if st.button("Login", key="hero_login", use_container_width=True):
                    st.session_state.page = 'login'
                    st.rerun()

    with hero_r:
        pill = (
            f'background:rgba(255,255,255,0.11);border:1px solid rgba(255,255,255,0.18);'
            f'border-radius:99px;padding:3px 10px;font-size:11px;color:{HM};margin-right:5px;'
        )
        # CSS keyframes — no f-string needed (no Python vars), so braces are literal
        _css = (
            '<style>'
            '@keyframes la-db{0%,80%,100%{transform:translateY(0)}40%{transform:translateY(-5px)}}'
            '@keyframes la-ut1{0%,4%{opacity:0;max-height:0;margin-bottom:0}5%,14%{opacity:1;max-height:44px;margin-bottom:6px}15%,100%{opacity:0;max-height:0;margin-bottom:0}}'
            '@keyframes la-um1{0%,14%{opacity:0;max-height:0;margin-bottom:0}16%,69%{opacity:1;max-height:60px;margin-bottom:6px}71%,100%{opacity:0;max-height:0;margin-bottom:0}}'
            '@keyframes la-bt1{0%,16%{opacity:0;max-height:0;margin-bottom:0}17%,26%{opacity:1;max-height:44px;margin-bottom:6px}28%,100%{opacity:0;max-height:0;margin-bottom:0}}'
            '@keyframes la-bm1{0%,27%{opacity:0;max-height:0;margin-bottom:0}29%,69%{opacity:1;max-height:220px;margin-bottom:6px}71%,100%{opacity:0;max-height:0;margin-bottom:0}}'
            '@keyframes la-ut2{0%,29%{opacity:0;max-height:0;margin-bottom:0}30%,36%{opacity:1;max-height:44px;margin-bottom:6px}38%,100%{opacity:0;max-height:0;margin-bottom:0}}'
            '@keyframes la-um2{0%,37%{opacity:0;max-height:0;margin-bottom:0}39%,69%{opacity:1;max-height:60px;margin-bottom:6px}71%,100%{opacity:0;max-height:0;margin-bottom:0}}'
            '@keyframes la-bt2{0%,39%{opacity:0;max-height:0;margin-bottom:0}40%,48%{opacity:1;max-height:44px;margin-bottom:6px}50%,100%{opacity:0;max-height:0;margin-bottom:0}}'
            '@keyframes la-bm2{0%,49%{opacity:0;max-height:0;margin-bottom:0}51%,69%{opacity:1;max-height:140px;margin-bottom:6px}71%,100%{opacity:0;max-height:0;margin-bottom:0}}'
            '.la-dot{display:inline-block;width:7px;height:7px;border-radius:50%;background:#94a3b8;animation:la-db 0.9s ease-in-out infinite}'
            '</style>'
        )
        # Typing dots — right (user) and left (bot)
        _dr = (
            '<div style="display:flex;justify-content:flex-end;">'
            '<div style="background:rgba(96,165,250,0.25);border:1px solid rgba(96,165,250,0.4);'
            'border-radius:12px 12px 2px 12px;padding:6px 11px;display:inline-flex;gap:4px;align-items:center;">'
            '<span class="la-dot" style="animation-delay:0s"></span>'
            '<span class="la-dot" style="animation-delay:0.15s"></span>'
            '<span class="la-dot" style="animation-delay:0.3s"></span>'
            '</div></div>'
        )
        _dl = (
            '<div style="display:flex;align-items:center;gap:6px;">'
            '<span style="font-size:16px;line-height:1;">&#9937;&#65039;</span>'
            '<div style="background:rgba(255,255,255,0.09);border:1px solid rgba(255,255,255,0.14);'
            'border-radius:12px 12px 12px 2px;padding:6px 11px;display:inline-flex;gap:4px;align-items:center;">'
            '<span class="la-dot" style="animation-delay:0s"></span>'
            '<span class="la-dot" style="animation-delay:0.15s"></span>'
            '<span class="la-dot" style="animation-delay:0.3s"></span>'
            '</div></div>'
        )
        st.markdown(
            _css
            # outer gradient wrapper
            + f'<div class="lifeaid-hero" style="background:{HB};border-radius:18px;padding:1.75rem 1.5rem;">'
            # chat card
            + f'<div style="background:rgba(10,18,40,0.7);border:1px solid rgba(255,255,255,0.12);border-radius:14px;overflow:hidden;">'
            # chat header
            + f'<div style="padding:0.8rem 1rem 0.7rem;border-bottom:1px solid rgba(255,255,255,0.1);display:flex;align-items:center;justify-content:space-between;">'
            + f'<span style="font-size:14px;font-weight:700;color:{HT};">&#9937;&#65039; LifeAid</span>'
            + f'<span style="font-size:11px;font-weight:600;color:#4ade80;">&#x25CF; Online</span>'
            + '</div>'
            # messages area
            + '<div style="padding:0.8rem 0.9rem 0.6rem;">'
            # ── user typing 1 ──
            + '<div style="overflow:hidden;animation:la-ut1 15s ease-in-out infinite;">' + _dr + '</div>'
            # ── user message 1 ──
            + f'<div style="overflow:hidden;animation:la-um1 15s ease-in-out infinite;">'
            + f'<div style="display:flex;justify-content:flex-end;">'
            + f'<div style="background:rgba(96,165,250,0.2);border:1px solid rgba(96,165,250,0.38);'
            + f'border-radius:12px 12px 2px 12px;padding:6px 11px;font-size:12px;color:{HT};max-width:88%;">'
            + 'Someone is choking!</div></div></div>'
            # ── bot typing 1 ──
            + '<div style="overflow:hidden;animation:la-bt1 15s ease-in-out infinite;">' + _dl + '</div>'
            # ── bot message 1 (CRITICAL) ──
            + f'<div style="overflow:hidden;animation:la-bm1 15s ease-in-out infinite;">'
            + f'<div style="display:flex;align-items:flex-start;gap:5px;">'
            + f'<span style="font-size:16px;flex-shrink:0;padding-top:1px;">&#9937;&#65039;</span>'
            + f'<div><p style="margin:0 0 3px;font-size:10px;font-weight:700;color:#fca5a5;letter-spacing:0.06em;">&#128680; CRITICAL EMERGENCY</p>'
            + f'<div style="background:rgba(239,68,68,0.16);border:1px solid rgba(239,68,68,0.36);'
            + f'border-radius:12px 12px 12px 2px;padding:7px 10px;font-size:11px;color:{HT};line-height:1.6;">'
            + '&#128680; Act immediately!<br>1. Stand behind them<br>2. Give 5 firm back blows<br>'
            + '3. Perform Heimlich maneuver<br>4. Repeat until object is clear'
            + '</div></div></div></div>'
            # ── user typing 2 ──
            + '<div style="overflow:hidden;animation:la-ut2 15s ease-in-out infinite;">' + _dr + '</div>'
            # ── user message 2 ──
            + f'<div style="overflow:hidden;animation:la-um2 15s ease-in-out infinite;">'
            + f'<div style="display:flex;justify-content:flex-end;">'
            + f'<div style="background:rgba(96,165,250,0.2);border:1px solid rgba(96,165,250,0.38);'
            + f'border-radius:12px 12px 2px 12px;padding:6px 11px;font-size:12px;color:{HT};max-width:88%;">'
            + 'What if they go unconscious?</div></div></div>'
            # ── bot typing 2 ──
            + '<div style="overflow:hidden;animation:la-bt2 15s ease-in-out infinite;">' + _dl + '</div>'
            # ── bot message 2 ──
            + f'<div style="overflow:hidden;animation:la-bm2 15s ease-in-out infinite;">'
            + f'<div style="display:flex;align-items:flex-start;gap:5px;">'
            + f'<span style="font-size:16px;flex-shrink:0;padding-top:1px;">&#9937;&#65039;</span>'
            + f'<div style="background:rgba(255,255,255,0.09);border:1px solid rgba(255,255,255,0.14);'
            + f'border-radius:12px 12px 12px 2px;padding:7px 10px;font-size:11px;color:{HT};line-height:1.6;">'
            + 'Call 911 immediately and start CPR.<br>'
            + 'Give 30 chest compressions then 2 rescue breaths.<br>'
            + 'A timer has been started for you &#10084;&#65039;'
            + '</div></div></div>'
            + '</div>'   # close messages area
            + '</div>'   # close chat card
            # pills row
            + f'<div style="margin-top:0.85rem;">'
            + f'<span style="{pill}">&#128203; CPR</span>'
            + f'<span style="{pill}">&#128293; Burns</span>'
            + f'<span style="{pill}">&#127758; 15+ Languages</span>'
            + f'<span style="background:rgba(255,255,255,0.11);border:1px solid rgba(255,255,255,0.18);border-radius:99px;padding:3px 10px;font-size:11px;color:{HM};">&#9201;&#65039; Timers</span>'
            + '</div>'
            + '</div>',  # close outer card
            unsafe_allow_html=True,
        )

    # ── Stats row ──
    st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)
    s1, s2, s3, s4 = st.columns(4, gap="small")
    stats = [
        (s1, "15+",  "Languages",        _C['primary']),
        (s2, "300+", "First Aid Topics",  _C['success']),
        (s3, "3",    "Emergency Timers",  _C['warning']),
        (s4, "24/7", "Always Available",  _C['danger']),
    ]
    for col, num, label, color in stats:
        with col:
            st.markdown(f"""
<div style="background:{_C['card']};border:1px solid {_C['border']};border-radius:12px;padding:1.25rem;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.04);">
  <p style="margin:0;font-size:30px;font-weight:800;color:{color};letter-spacing:-0.03em;">{num}</p>
  <p style="margin:0.2rem 0 0;font-size:13px;color:{_C['muted']};font-weight:500;">{label}</p>
</div>
""", unsafe_allow_html=True)

    # ── Feature cards ──
    st.markdown(f'<h2 style="font-size:20px;font-weight:700;color:{_C["text"]};margin:1.75rem 0 1rem;letter-spacing:-0.02em;">Everything you need in an emergency</h2>', unsafe_allow_html=True)

    fc = st.columns(5, gap="medium")
    features = [
        ("📋", "Step-by-Step Instructions",  "Clear sequential guidance for any emergency"),
        ("🫀", "CPR & Choking Guidance",      "Lifesaving procedures with integrated timers"),
        ("📝", "Symptom Assessment",          "Severity: Critical, Urgent, Moderate"),
        ("🔥", "Wound & Burn Care",           "Wound treatment and cooling protocols"),
        ("💾", "Save Sessions",               "Log in to save and revisit your guides"),
    ]
    for col, (icon, title, desc) in zip(fc, features):
        with col:
            st.markdown(f"""
<div style="background:{_C['card']};border:1px solid {_C['border']};border-radius:14px;padding:1.4rem 1rem;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.05);min-height:200px;height:100%;display:flex;flex-direction:column;justify-content:center;align-items:center;">
  <div style="font-size:34px;margin-bottom:0.65rem;">{icon}</div>
  <p style="margin:0 0 0.4rem;font-size:13px;font-weight:700;color:{_C['text']};line-height:1.35;">{title}</p>
  <p style="margin:0;font-size:11px;color:{_C['muted']};line-height:1.5;">{desc}</p>
</div>
""", unsafe_allow_html=True)

    # ── Disclaimer ──
    st.markdown(f"""
<div style="background:{_C['card']};border:1px solid {_C['border']};border-left:4px solid {_C['primary']};border-radius:10px;padding:1rem 1.25rem;margin:1.75rem 0 1.25rem;display:table;width:100%;">
  <div style="display:table-cell;vertical-align:middle;font-size:18px;width:28px;padding-right:12px;">ℹ️</div>
  <div style="display:table-cell;vertical-align:middle;">
    <p style="margin:0;font-size:13px;color:{_C['muted']};line-height:1.55;">
      <strong style="color:{_C['text']};">Medical Disclaimer:</strong>
      LifeAid is an AI assistant for first aid guidance only and is not a replacement for professional medical care.
      Always call emergency services (911) for life-threatening situations.
    </p>
  </div>
</div>
""", unsafe_allow_html=True)

    # ── Footer ──
    st.markdown(f"""
<div style="border-top:1px solid {_C['border']};padding:1rem 0;display:table;width:100%;">
  <div style="display:table-cell;vertical-align:middle;font-size:12px;color:{_C['muted']};">&#169; 2025 LifeAid Assistant</div>
  <div style="display:table-cell;vertical-align:middle;text-align:right;font-size:12px;">
    <a href="#" style="color:{_C['muted']};text-decoration:none;margin-left:1.5rem;">Privacy Policy</a>
    <a href="#" style="color:{_C['muted']};text-decoration:none;margin-left:1.5rem;">Terms of Use</a>
    <a href="#" style="color:{_C['muted']};text-decoration:none;margin-left:1.5rem;">Contact</a>
  </div>
</div>
""", unsafe_allow_html=True)


# ─── Auth pages ───────────────────────────────────────────────────────────────
def _auth_page_wrapper(title, icon, subtitle):
    """Centered card layout for auth pages."""
    st.markdown(f"""
    <div style="min-height:10vh; display:flex; align-items:flex-start;
                justify-content:center; padding-top:3rem;">
    </div>
    """, unsafe_allow_html=True)
    _, center, _ = st.columns([1.5, 2, 1.5])
    with center:
        st.markdown(f"""
        <div style="text-align:center; margin-bottom:1.75rem; animation:fadeUp 0.4s ease;">
            <div style="font-size:52px; line-height:1; margin-bottom:0.75rem;">{icon}</div>
            <h1 style="font-size:28px; font-weight:800; color:{_C['text']};
                       margin:0; letter-spacing:-0.03em;">{title}</h1>
            <p style="margin:0.4rem 0 0; font-size:14px; color:{_C['muted']};">{subtitle}</p>
        </div>
        """, unsafe_allow_html=True)
    return center


def show_login_page():
    center = _auth_page_wrapper("Welcome back", "⛑️", "Sign in to your LifeAid account")
    with center:
        with st.form("login_form", clear_on_submit=False):
            username = st.text_input("Username", placeholder="Enter your username")
            password = st.text_input("Password", type="password",
                                     placeholder="Enter your password")
            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
            la, lb = st.columns(2)
            with la:
                login_btn = st.form_submit_button("🔓 Sign In", use_container_width=True)
            with lb:
                signup_link = st.form_submit_button("📝 Create Account",
                                                    use_container_width=True)
            if login_btn:
                if username and password:
                    ok, msg = auth.login(username, password)
                    if ok:
                        st.session_state.logged_in = True
                        st.session_state.username = username
                        st.session_state.page = 'chat'
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
                else:
                    st.warning("Please enter both username and password.")
            if signup_link:
                st.session_state.page = 'signup'
                st.rerun()

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
        if st.button("← Back to Home", use_container_width=True, key="login_back"):
            st.session_state.page = 'landing'
            st.rerun()


def show_signup_page():
    center = _auth_page_wrapper("Create your account", "⛑️",
                                "Join LifeAid to save your emergency sessions")
    with center:
        with st.form("signup_form", clear_on_submit=False):
            username = st.text_input("Username", placeholder="At least 3 characters")
            password = st.text_input("Password", type="password",
                                     placeholder="At least 6 characters")
            confirm = st.text_input("Confirm Password", type="password",
                                    placeholder="Re-enter your password")
            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
            sa, sb = st.columns(2)
            with sa:
                signup_btn = st.form_submit_button("✅ Create Account",
                                                   use_container_width=True)
            with sb:
                login_link = st.form_submit_button("🔓 Sign In Instead",
                                                   use_container_width=True)
            if signup_btn:
                if username and password and confirm:
                    if auth.user_exists(username):
                        st.error("Username already taken.")
                    elif password != confirm:
                        st.error("Passwords don't match.")
                    elif len(username) < 3:
                        st.error("Username must be at least 3 characters.")
                    elif len(password) < 6:
                        st.error("Password must be at least 6 characters.")
                    else:
                        ok, msg = auth.signup(username, password)
                        if ok:
                            st.success("Account created! You can now sign in.")
                            st.session_state.page = 'login'
                            st.balloons()
                            st.rerun()
                        else:
                            st.error(msg)
                else:
                    st.warning("Please fill in all fields.")
            if login_link:
                st.session_state.page = 'login'
                st.rerun()

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
        if st.button("← Back to Home", use_container_width=True, key="signup_back"):
            st.session_state.page = 'landing'
            st.rerun()


# ─── Chat page ────────────────────────────────────────────────────────────────
def show_chat_page():
    # ── CSS: dark panel for the left (fake-sidebar) column only ───────────────
    st.markdown("""
    <style>
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:first-child {
        background: #1a1a2e !important;
        border-radius: 12px !important;
        border: 1px solid #e63946 !important;
        padding: 0.5rem 0.6rem 0.75rem !important;
    }
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:first-child p,
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:first-child span,
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:first-child h3,
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:first-child label,
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:first-child small {
        color: #f0f0f0 !important;
    }
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:first-child .stButton > button {
        background: rgba(255,255,255,0.08) !important;
        border: 1px solid rgba(255,255,255,0.22) !important;
        color: #f0f0f0 !important;
        text-align: left !important;
        border-radius: 6px !important;
        font-size: 13px !important;
    }
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:first-child .stButton > button:hover {
        background: rgba(255,255,255,0.18) !important;
        color: #ffffff !important;
    }
    /* Circular settings button (targeted via its tooltip title attribute) */
    button[title="Settings"] {
        border-radius: 50% !important;
        width: 40px !important;
        height: 40px !important;
        padding: 0 !important;
        min-width: 0 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Chat panel card styling ────────────────────────────────────────────────
    _chat_card_bg     = '#161B22' if _DK else '#FFFFFF'
    _chat_card_border = '#30363D' if _DK else '#C8C4BC'
    _chat_card_shadow = ('0 4px 24px rgba(0,0,0,0.40), 0 1px 4px rgba(0,0,0,0.25)'
                         if _DK else
                         '0 4px 24px rgba(0,0,0,0.10), 0 1px 4px rgba(0,0,0,0.06)')
    st.markdown(f"""
    <style>
    /* Right panel (chat area) — card/panel frame */
    .main .block-container
      > div[data-testid="stVerticalBlock"]
      > div[data-testid="stHorizontalBlock"]
      > div[data-testid="column"]:last-child {{
        background: {_chat_card_bg} !important;
        border: 1.5px solid {_chat_card_border} !important;
        border-radius: 14px !important;
        box-shadow: {_chat_card_shadow} !important;
        padding: 1.1rem 1.4rem 1.25rem !important;
    }}
    </style>
    """, unsafe_allow_html=True)

    left_col, right_col = st.columns([1, 4])

    # ── Left panel (fake sidebar) ──────────────────────────────────────────────
    with left_col:
        st.markdown("### ⛑️ LifeAid")

        # ── New Chat + Settings row ──
        _nc, _sc = st.columns([3, 1])
        with _nc:
            if st.button("➕ New Chat", use_container_width=True, key="sb_new_chat"):
                st.session_state.chat_history = []
                st.session_state.chatbot.reset_conversation()
                st.session_state.active_timer = None
                st.session_state.timer_start_time = None
                st.session_state.timer_paused = False
                st.session_state.timer_paused_remaining = None
                st.session_state.completion_sound_played = False
                st.session_state.current_chat_saved = False
                if hasattr(st.session_state.chatbot, 'chatbot'):
                    cb = st.session_state.chatbot.chatbot
                    if hasattr(cb, 'last_matched_emergency'):
                        cb.last_matched_emergency = None
                    if hasattr(cb, 'conversation_history'):
                        cb.conversation_history = []
                st.rerun()
        with _sc:
            if st.button("⚙️", key="sb_settings", help="Settings"):
                st.session_state.show_settings = not st.session_state.show_settings
                st.rerun()

        # ── Settings panel ──
        if st.session_state.show_settings:
            st.markdown("""
            <div style="background:rgba(255,255,255,0.05); border-radius:10px;
                        padding:1rem; border:1px solid rgba(255,255,255,0.15);
                        margin:0.5rem 0;">
            """, unsafe_allow_html=True)

            # Section 1 – Theme
            theme_choice = st.radio(
                "🎨 Theme",
                ["☀️ Light", "🌙 Dark"],
                index=1 if st.session_state.dark_mode else 0,
                key="theme_radio",
            )
            if theme_choice == "🌙 Dark" and not st.session_state.dark_mode:
                st.session_state.dark_mode = True
                st.rerun()
            elif theme_choice == "☀️ Light" and st.session_state.dark_mode:
                st.session_state.dark_mode = False
                st.rerun()

            st.markdown("---")

            # Section 2 – Feedback
            st.markdown("**📝 Feedback**")

            def _reset_feedback_flag():
                st.session_state.feedback_submitted = False

            feedback_text = st.text_area(
                "Share your feedback",
                key=f"feedback_input_{st.session_state.feedback_key}",
                label_visibility="collapsed",
                placeholder="Type your feedback here…",
                on_change=_reset_feedback_flag,
            )
            if st.button("Submit", key="submit_feedback", use_container_width=True):
                if feedback_text.strip():
                    st.session_state.feedback_list.append(feedback_text.strip())
                    st.session_state.feedback_key += 1
                    st.session_state.feedback_submitted = True
                    st.rerun()
                else:
                    st.warning("Please enter some feedback first.")
            if st.session_state.feedback_submitted:
                st.success("Thank you for your feedback!")

            st.markdown("---")

            # Section 3 – Help
            st.markdown("**❓ Help**")
            st.caption("Help content coming soon...")

            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("**⚡ Quick Help**")
        if st.button("🩸 Bleeding", use_container_width=True, key="sb_bleeding"):
            handle_quick_help('bleeding')
        if st.button("🫁 Choking", use_container_width=True, key="sb_choking"):
            handle_quick_help('choking')
        if st.button("❤️ CPR", use_container_width=True, key="sb_cpr"):
            handle_quick_help('cpr')
        if st.button("🔥 Burns", use_container_width=True, key="sb_burns"):
            handle_quick_help('burns')

        st.markdown("---")
        st.markdown("**💬 Saved Chats**")
        if st.session_state.logged_in:
            saved_chats = auth.get_user_chats(st.session_state.username)
            if saved_chats:
                for idx, chat_session in enumerate(reversed(saved_chats)):
                    msgs = chat_session['messages']
                    preview = "Chat " + str(idx + 1)
                    for m in msgs:
                        if m['role'] == 'user':
                            raw = m['content']
                            preview = (raw[:22] + "…" if len(raw) > 22 else raw)
                            break
                    actual_idx = len(saved_chats) - 1 - idx
                    c1, c2 = st.columns([4, 1])
                    with c1:
                        if st.button(preview, key=f"load_{idx}",
                                     use_container_width=True):
                            st.session_state.chat_history = saved_chats[actual_idx]['messages']
                            st.session_state.page = 'chat'
                            st.rerun()
                    with c2:
                        if st.button("×", key=f"del_{idx}"):
                            auth.delete_user_chat(st.session_state.username, actual_idx)
                            st.rerun()
            else:
                st.caption("No saved chats yet.")
        else:
            st.caption("🔒 Login to save chats")

    # ── Right panel (chat content) ─────────────────────────────────────────────
    with right_col:
        # ── Header bar ──
        hc1, hc2, hc3, hc4 = st.columns([1.2, 6, 0.8, 1.2])
        with hc1:
            if st.button("← Home", key="chat_back"):
                st.session_state.page = 'landing'
                st.rerun()
        with hc2:
            if st.session_state.logged_in:
                st.markdown(f"""
                <p style="margin:0; font-size:15px; font-weight:600; color:{_C['text']};
                          padding-top:0.45rem;">
                    ⛑️ LifeAid &nbsp;·&nbsp;
                    <span style="font-weight:400; color:{_C['muted']};">
                        Welcome back, <strong style="color:{_C['primary']};">
                        {st.session_state.username}</strong>
                    </span>
                </p>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <p style="margin:0; font-size:18px; font-weight:600; color:{_C['text']};
                          padding-top:0.45rem;">
                    ⛑️ LifeAid &nbsp;
                    <span style="font-weight:400; font-size:15px; color:{_C['muted']};">
                        — Emergency First Aid Assistant
                    </span>
                </p>""", unsafe_allow_html=True)
        with hc3:
            if st.button("🗑️", key="clear_chat_btn", help="Clear chat"):
                st.session_state.chat_history = []
                st.session_state.chatbot.reset_conversation()
                st.session_state.active_timer = None
                st.session_state.timer_start_time = None
                st.session_state.timer_paused = False
                st.session_state.timer_paused_remaining = None
                st.session_state.completion_sound_played = False
                st.session_state.current_chat_saved = False
                if hasattr(st.session_state.chatbot, 'chatbot'):
                    cb = st.session_state.chatbot.chatbot
                    if hasattr(cb, 'last_matched_emergency'):
                        cb.last_matched_emergency = None
                    if hasattr(cb, 'conversation_history'):
                        cb.conversation_history = []
                st.rerun()
        with hc4:
            if st.session_state.logged_in:
                if st.button(" Logout", key="chat_logout_btn", use_container_width=True):
                    st.session_state.logged_in = False
                    st.session_state.username = None
                    st.session_state.chat_history = []
                    st.session_state.current_chat_saved = False
                    st.rerun()
            else:
                if st.button(" Login", key="chat_login_btn", use_container_width=True):
                    st.session_state.page = 'login'
                    st.rerun()

        st.markdown(f"<hr style='margin:0.5rem 0 1rem;border-color:{_C['border']};'>",
                    unsafe_allow_html=True)
    
        # ── Timer ──
        st.markdown('<div id="timer-section"></div>', unsafe_allow_html=True)
    
        if st.session_state.active_timer:
            timer_info = st.session_state.active_timer
    
            if st.session_state.timer_paused and st.session_state.timer_paused_remaining is not None:
                remaining = st.session_state.timer_paused_remaining
            elif st.session_state.timer_start_time:
                elapsed = (datetime.datetime.now() - st.session_state.timer_start_time).total_seconds()
                remaining = max(0, timer_info['duration'] - elapsed)
            else:
                remaining = timer_info['duration']
    
            if remaining > 0:
                mins, secs = divmod(int(remaining), 60)
                paused_label = " &nbsp;⏸ PAUSED" if st.session_state.timer_paused else ""
                pct = max(0.0, min(1.0, (timer_info['duration'] - remaining) / timer_info['duration']))
    
                pulse_style = "" if st.session_state.timer_paused else "animation:timerGlow 2s infinite;"
                st.markdown(f"""
                <div style="background:linear-gradient(135deg, #1D4ED8 0%, #4F46E5 100%);
                            border-radius:16px; padding:1.4rem 1.75rem;
                            margin-bottom:1.25rem; {pulse_style}
                            box-shadow:0 8px 32px rgba(79,70,229,0.35);
                            border:1px solid rgba(255,255,255,0.15);">
                    <div style="display:flex; align-items:center; justify-content:space-between;
                                flex-wrap:wrap; gap:0.5rem;">
                        <div>
                            <p style="margin:0; font-size:12px; font-weight:700;
                                      color:rgba(255,255,255,0.6); letter-spacing:0.08em;
                                      text-transform:uppercase;">
                                ACTIVE TIMER{paused_label}
                            </p>
                            <h3 style="margin:0.15rem 0 0; font-size:20px; color:white;
                                       font-weight:700;">{timer_info['title']}</h3>
                            <p style="margin:0.2rem 0 0; font-size:13px;
                                      color:rgba(255,255,255,0.7);">
                                {timer_info['instructions']}
                            </p>
                        </div>
                        <div style="text-align:right;">
                            <p style="margin:0; font-size:52px; font-weight:800;
                                      color:white; letter-spacing:-0.02em;
                                      font-variant-numeric:tabular-nums;
                                      line-height:1;">
                                {mins:02d}:{secs:02d}
                            </p>
                            <a href="#chat-section" style="font-size:11px; color:rgba(255,255,255,0.6);
                               text-decoration:none;">↓ Go to chat</a>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
    
                if not st.session_state.timer_paused and 'CPR' in timer_info['title']:
                    st.markdown("""
                    <audio autoplay loop>
                        <source src="https://actions.google.com/sounds/v1/alarms/beep_short.ogg"
                                type="audio/ogg">
                    </audio>
                    """, unsafe_allow_html=True)
    
                st.session_state.completion_sound_played = False
                st.progress(pct)
    
                tc1, tc2, tc3 = st.columns(3)
                with tc1:
                    if st.session_state.timer_paused:
                        if st.button("▶️ Resume", key="resume_timer", use_container_width=True):
                            rem = st.session_state.timer_paused_remaining
                            st.session_state.timer_start_time = (
                                datetime.datetime.now()
                                - datetime.timedelta(seconds=(timer_info['duration'] - rem))
                            )
                            st.session_state.timer_paused = False
                            st.session_state.timer_paused_remaining = None
                            st.rerun()
                    else:
                        if st.button("⏸️ Pause", key="pause_timer", use_container_width=True):
                            st.session_state.timer_paused = True
                            st.session_state.timer_paused_remaining = remaining
                            st.rerun()
                with tc2:
                    if st.button("🔄 Restart", key="restart_timer", use_container_width=True):
                        st.session_state.timer_start_time = datetime.datetime.now()
                        st.session_state.timer_paused = False
                        st.session_state.timer_paused_remaining = None
                        st.rerun()
                with tc3:
                    if st.button("⏹️ Stop", key="stop_timer", use_container_width=True):
                        st.session_state.active_timer = None
                        st.session_state.timer_start_time = None
                        st.session_state.timer_paused = False
                        st.session_state.timer_paused_remaining = None
                        st.session_state.completion_sound_played = False
                        st.rerun()
    
                st.markdown(f"<hr style='border-color:{_C['border']};margin:1rem 0;'>",
                            unsafe_allow_html=True)
    
            else:
                st.success(f"✅ {timer_info['title']} Complete!")
                if not st.session_state.completion_sound_played:
                    st.markdown("""
                    <audio autoplay>
                        <source src="https://actions.google.com/sounds/v1/alarms/beep_short.ogg"
                                type="audio/ogg">
                    </audio>""" * 3, unsafe_allow_html=True)
                    st.session_state.completion_sound_played = True
                cc1, cc2 = st.columns(2)
                with cc1:
                    if st.button("🔄 Restart Timer", key="restart_complete_btn",
                                 use_container_width=True):
                        st.session_state.timer_start_time = datetime.datetime.now()
                        st.session_state.timer_paused = True
                        st.session_state.timer_paused_remaining = timer_info['duration']
                        st.session_state.completion_sound_played = False
                        st.rerun()
                with cc2:
                    if st.button("✓ Clear Timer", key="clear_timer_btn", use_container_width=True):
                        st.session_state.active_timer = None
                        st.session_state.timer_start_time = None
                        st.session_state.timer_paused = False
                        st.session_state.timer_paused_remaining = None
                        st.session_state.completion_sound_played = False
                        st.rerun()
    
        # ── Chat messages ──
        st.markdown('<div id="chat-section"></div>', unsafe_allow_html=True)
    
        if not st.session_state.chat_history:
            st.markdown(f"""
            <div style="background:{_C['primary']}10; border:1px solid {_C['primary']}28;
                        border-radius:14px; padding:1.25rem 1.5rem; margin:0.5rem 0 1rem;
                        display:flex; align-items:flex-start; gap:0.85rem;">
                <span style="font-size:24px; flex-shrink:0;">👋</span>
                <div>
                    <p style="margin:0; font-size:15px; font-weight:600; color:{_C['text']};">
                        Hello! I'm LifeAid.
                    </p>
                    <p style="margin:0.25rem 0 0; font-size:13px; color:{_C['muted']};
                              line-height:1.55;">
                        Describe your emergency and I'll guide you through first aid steps.
                        Use the Quick Help buttons in the sidebar for instant guidance.
                    </p>
                </div>
            </div>
            """, unsafe_allow_html=True)
    
        import hashlib
    
        for idx, message in enumerate(st.session_state.chat_history):
            if message['role'] == 'user':
                _usr_bg  = '#1F6FEB' if _DK else '#0969DA'
                _usr_txt = '#FFFFFF'
                st.markdown(
                    f"<div style='display:flex; justify-content:flex-end;"
                    f" align-items:flex-end; gap:8px; margin:4px 0;'>"
                    f"<div style='background:{_usr_bg}; color:{_usr_txt};"
                    f" padding:0.65rem 1rem; border-radius:18px 18px 4px 18px;"
                    f" max-width:82%; word-wrap:break-word; overflow-wrap:break-word;"
                    f" font-size:14px; line-height:1.65;"
                    f" box-shadow:0 2px 8px rgba(0,0,0,0.20);'>"
                    f"{message['content']}</div>"
                    f"<div style='font-size:22px; flex-shrink:0; line-height:1;'>👤</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            else:
                with st.chat_message("assistant", avatar="⛑️"):
                    content = clean_response(message['content'])
                    is_arabic = any('\u0600' <= c <= '\u06FF' for c in content)
                    direction = "rtl" if is_arabic else "ltr"
                    text_align = "right" if is_arabic else "left"
                    if content.startswith('🚨'):
                        severity = 'critical'
                    elif content.startswith('⚠️'):
                        severity = 'urgent'
                    elif content.startswith('🟢'):
                        severity = 'moderate'
                    else:
                        severity = 'normal'
                    formatted = content.replace('\n\n', '<br><br>').replace('\n', '<br>')
    
                    content_hash = hashlib.md5(content.encode()).hexdigest()[:8]
                    message_id = f"msg_{idx}_{content_hash}"
    
                    # Evict old audio caches
                    _audio_keys = [k for k in st.session_state.keys()
                                    if k.startswith('audio_cache_msg_')]
                    if len(_audio_keys) > 15:
                        def _msg_idx(k):
                            try: return int(k.split('_')[3])
                            except: return 0
                        for _old in sorted(_audio_keys, key=_msg_idx)[:len(_audio_keys) - 15]:
                            del st.session_state[_old]
    
                    sev_styles = {
                        'critical': (
                            f"background:{'#2D0A0A' if _DK else '#7F1D1D'};"
                            f"border:1px solid {'#F85149' if _DK else '#F85149'};"
                            f"border-left:4px solid {'#F85149' if _DK else '#EF4444'};"
                            f"color:{'#FCA5A5' if _DK else '#FCA5A5'};"
                            f"box-shadow:0 4px 16px rgba(239,68,68,0.35);"
                        ),
                        'urgent': (
                            f"background:{'#2D1B00' if _DK else '#78350F'};"
                            f"border:1px solid {'#FFA657' if _DK else '#FFA657'};"
                            f"border-left:4px solid {'#FFA657' if _DK else '#F59E0B'};"
                            f"color:{'#FFA657' if _DK else '#FED7AA'};"
                            f"box-shadow:0 4px 16px rgba(245,158,11,0.35);"
                        ),
                        'moderate': (
                            f"background:{'#052E16' if _DK else '#14532D'};"
                            f"border:1px solid {'#56D364' if _DK else '#56D364'};"
                            f"border-left:4px solid {'#56D364' if _DK else '#22C55E'};"
                            f"color:{'#86EFAC' if _DK else '#86EFAC'};"
                            f"box-shadow:0 4px 16px rgba(34,197,94,0.35);"
                        ),
                        'normal': (
                            f"background:{'#2D333B' if _DK else '#ECEAE4'};"
                            f"border:1px solid {'#444C56' if _DK else '#D4D0C8'};"
                            f"color:{'#E6EDF3' if _DK else '#1E293B'};"
                            f"box-shadow:0 2px 8px rgba(0,0,0,{'0.20' if _DK else '0.07'});"
                        ),
                    }
                    style = sev_styles.get(severity, sev_styles['normal'])
                    _radius = '18px 18px 18px 4px' if severity == 'normal' else '12px'

                    st.markdown(f"""
                    <div style="{style}
                                 padding:1rem 1.25rem; border-radius:{_radius};
                                 word-wrap:break-word; overflow-wrap:break-word;
                                 line-height:1.75; font-size:14px;
                                 direction:{direction}; text-align:{text_align};">
                        {formatted}
                    </div>
                    """, unsafe_allow_html=True)
                    text_to_speech_button(content, message_id)
    
        # ── Emergency numbers button + panel ──
        st.markdown("""
        <style>
        @keyframes sos-pulse {
            0%, 100% { box-shadow: 0 0 0 0 rgba(220, 38, 38, 0.7),
                                   0 0 0 0 rgba(220, 38, 38, 0.4); }
            50%       { box-shadow: 0 0 0 6px rgba(220, 38, 38, 0.2),
                                   0 0 0 12px rgba(220, 38, 38, 0); }
        }
        button[title="Emergency Numbers"] {
            background: linear-gradient(135deg, #dc2626 0%, #991b1b 100%) !important;
            border: 1.5px solid #ef4444 !important;
            border-radius: 8px !important;
            font-size: 20px !important;
            font-weight: 700 !important;
            color: #ffffff !important;
            letter-spacing: 0.03em !important;
            animation: sos-pulse 2s ease-in-out infinite !important;
            transition: all 0.18s ease !important;
            box-shadow: 0 4px 14px rgba(220, 38, 38, 0.55),
                        inset 0 1px 0 rgba(255,255,255,0.15) !important;
            text-shadow: 0 1px 2px rgba(0,0,0,0.4) !important;
        }
        button[title="Emergency Numbers"]:hover {
            background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%) !important;
            border-color: #fca5a5 !important;
            box-shadow: 0 6px 20px rgba(220, 38, 38, 0.75),
                        inset 0 1px 0 rgba(255,255,255,0.2) !important;
            transform: translateY(-2px) scale(1.04) !important;
            animation: none !important;
        }
        button[title="Emergency Numbers"]:active {
            transform: translateY(0) scale(0.97) !important;
            box-shadow: 0 2px 8px rgba(220, 38, 38, 0.5) !important;
        }
        </style>
        """, unsafe_allow_html=True)

        # ── Emergency numbers panel (shows above the input row when open) ──
        if st.session_state.show_emergency_numbers:
            st.markdown("""
<div style="background:#7f1d1d; border:1px solid #ef4444;
            border-radius:12px; padding:1rem 1.25rem;
            margin-bottom:0.5rem;">
    <p style="color:#fca5a5; font-weight:700; font-size:13px;
              margin:0 0 0.6rem;">🆘 Emergency Numbers</p>
    <div style="display:grid; grid-template-columns:1fr 1fr; gap:0.4rem;">
        <div style="color:white; font-size:12px;">🇺🇸 USA / Canada</div>
        <div style="color:#fca5a5; font-size:13px; font-weight:700;">911</div>
        <div style="color:white; font-size:12px;">🇬🇧 UK</div>
        <div style="color:#fca5a5; font-size:13px; font-weight:700;">999</div>
        <div style="color:white; font-size:12px;">🇪🇺 Europe</div>
        <div style="color:#fca5a5; font-size:13px; font-weight:700;">112</div>
        <div style="color:white; font-size:12px;">🇦🇺 Australia</div>
        <div style="color:#fca5a5; font-size:13px; font-weight:700;">000</div>
        <div style="color:white; font-size:12px;">🇸🇦 Saudi Arabia</div>
        <div style="color:#fca5a5; font-size:13px; font-weight:700;">911</div>
        <div style="color:white; font-size:12px;">🇪🇬 Egypt</div>
        <div style="color:#fca5a5; font-size:13px; font-weight:700;">123</div>
        <div style="color:white; font-size:12px;">🇦🇪 UAE</div>
        <div style="color:#fca5a5; font-size:13px; font-weight:700;">998</div>
        <div style="color:white; font-size:12px;">🇯🇴 Jordan</div>
        <div style="color:#fca5a5; font-size:13px; font-weight:700;">911</div>
    </div>
    <p style="color:rgba(255,255,255,0.5); font-size:10px;
              margin:0.6rem 0 0;">
        Tap a number on your phone to call
    </p>
</div>
""", unsafe_allow_html=True)

        # ── Eval download panel (shown only after secret keyword) ────────
        if st.session_state.get('show_eval_download'):
            if not _interactions:
                st.info("No evaluation data yet")
                st.session_state.show_eval_download = False
            else:
                _fname = f"evaluation_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                _dl_col, _x_col = st.columns([5, 1])
                with _dl_col:
                    st.download_button(
                        label="Download Evaluation Report",
                        data=_build_eval_excel(),
                        file_name=_fname,
                        mime=(
                            "application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet"
                        ),
                        key="eval_download_btn",
                        use_container_width=True,
                    )
                with _x_col:
                    if st.button("Close", key="close_eval_download",
                                 use_container_width=True):
                        st.session_state.show_eval_download = False
                        st.rerun()

        # ── Chat input row: SOS button left, chat input right ──
        sos_col, input_col = st.columns([1, 12])
        with sos_col:
            if st.button("📞", key="sos_call_btn", help="Emergency Numbers",
                         use_container_width=True):
                st.session_state.show_emergency_numbers = not st.session_state.get(
                    'show_emergency_numbers', False)
                st.rerun()
        with input_col:
            user_input = st.chat_input("Describe the emergency…", key="chat_input")
    
        if user_input:
            if user_input.strip() == "evaldownload123":
                st.session_state.show_eval_download = True
                st.rerun()

            st.session_state.skip_refresh = True
            st.session_state.chat_history.append({'role': 'user', 'content': user_input})

            # Capture whether we are mid-follow-up BEFORE calling get_response.
            # Used below to prevent timer detection misfiring on follow-up answers.
            _was_in_followup = (
                hasattr(st.session_state.chatbot, 'chatbot') and
                st.session_state.chatbot.chatbot.conversation_state.get(
                    'waiting_for_followup', False)
            )

            with st.spinner('💭 Thinking...'):
                _t0 = time.perf_counter()
                _eval_error = None
                try:
                    bot_response = st.session_state.chatbot.get_response(user_input)
                except Exception as _exc:
                    bot_response = ""
                    _eval_error = str(_exc)
                _eval_time = time.perf_counter() - _t0
                matched_severity = None
                if hasattr(st.session_state.chatbot, 'chatbot'):
                    matched_data = getattr(st.session_state.chatbot.chatbot,
                                           'last_matched_emergency', None)
                    if matched_data and 'severity' in matched_data:
                        matched_severity = matched_data['severity']

            if not bot_response:
                bot_response = "I'm here to help. Please describe your emergency."
    
            user_lower = user_input.lower()
            new_q_kws = ['what', 'how', 'when', 'where', 'why', 'should', 'do i',
                         'ماذا', 'كيف', 'هل', 'متى', 'que', 'como', 'qué', 'comment', 'quoi']
            emerg_kws = ['bleeding', 'blood', 'cardiac', 'choking', 'burn', 'fracture', 'poison',
                         'unconscious', 'seizure', 'stroke', 'heart attack', 'cpr',
                         'نزيف', 'دم', 'قلب', 'اختناق', 'حرق',
                         'saignement', 'cardiaque', 'brûlure', 'sangrado', 'cardíaco', 'quemadura']
    
            is_new_question = (
                '?' in user_input
                or any(w in user_lower for w in new_q_kws)
                or any(w in user_lower for w in emerg_kws)
                or len(user_input.split()) > 8
            )
            if not is_new_question:
                short_answers = ['yes', 'no', 'yeah', 'nope', 'sure', 'ok', 'okay', 'نعم', 'لا']
                if not (any(user_lower.startswith(a) for a in short_answers)
                        or len(user_input.split()) < 4):
                    is_new_question = True
    
            response_upper = bot_response.upper()
            if matched_severity:
                sev_map = {'CRITICAL': 'critical', 'URGENT': 'urgent',
                           'MODERATE': 'moderate', 'MILD': 'moderate'}
                severity = sev_map.get(matched_severity.upper(), 'normal')
                if severity == 'moderate' and '🟢 MODERATE' not in bot_response:
                    bot_response = '🟢 MODERATE\n\n' + bot_response
            elif '🚨 CRITICAL' in bot_response or 'CRITICAL EMERGENCY' in response_upper:
                severity = 'critical'
            elif '⚠️ URGENT' in bot_response or 'URGENT' in response_upper:
                severity = 'urgent'
            elif 'MODERATE' in response_upper:
                severity = 'moderate'
                if '🟢 MODERATE' not in bot_response:
                    bot_response = '🟢 MODERATE\n\n' + bot_response
            else:
                severity = 'normal'

            # If the response contains none of the severity markers it is a
            # plain conversational reply (thanks, bye, ok, etc.).  Override
            # any severity that was carried over from a previous turn.
            _has_sev_marker = (
                '🚨' in bot_response or '⚠️' in bot_response or '🟢' in bot_response
                or 'CRITICAL' in response_upper or 'URGENT' in response_upper
                or 'MODERATE' in response_upper
            )
            if not _has_sev_marker:
                severity = 'normal'
    
            st.session_state.chat_history.append(
                {'role': 'assistant', 'content': bot_response, 'severity': severity})

            # ── Evaluator: record this interaction ────────────────────────
            _detected_lang = getattr(st.session_state.chatbot, 'user_language', 'en')
            _match_found = (
                matched_severity is not None
                or ('CRITICAL' in bot_response.upper() or 'URGENT' in bot_response.upper()
                    or severity in ('critical', 'urgent', 'moderate'))
            ) and 'not sure' not in bot_response.lower() and "couldn't find" not in bot_response.lower()
            record_interaction(
                user_input=user_input,
                bot_response=bot_response,
                response_time=_eval_time,
                severity=severity,
                match_found=_match_found,
                detected_lang=_detected_lang,
                input_lang=_detected_lang,
                error=_eval_error,
            )

            if st.session_state.logged_in and len(st.session_state.chat_history) >= 2:
                if not st.session_state.current_chat_saved:
                    auth.save_user_chat(st.session_state.username, st.session_state.chat_history)
                    st.session_state.current_chat_saved = True
                else:
                    auth.update_user_chat(st.session_state.username, st.session_state.chat_history)
    
            if is_new_question:
                # When the user was answering a follow-up question, skip
                # scanning the bot response for timer keywords — the response
                # is a specific follow-up reply, not a fresh emergency, and
                # topic words in it (e.g. "cool", "water") would otherwise
                # trigger the wrong timer.  The keyword fallback on user_input
                # (ulq) below still runs so a user who genuinely introduces a
                # new emergency mid-conversation still gets the correct timer.
                timer_needed = (
                    None if _was_in_followup else detect_timer_need(bot_response, user_input)
                )
                if not timer_needed:
                    ulq = user_input.lower()
                    # Bleeding questions must never trigger a timer via the fallback path either.
                    _fb_bleed = any(w in ulq for w in ['bleed', 'bleeding', 'blood loss', 'hemorrhage', 'haemorrhage'])
                    _fb_override = any(w in ulq for w in ['cpr', 'cardiac', 'heart attack', 'burn', 'chok', 'heimlich'])
                    if not (_fb_bleed and not _fb_override):
                        if any(w in ulq for w in ['cpr', 'cardiac', 'heart attack', 'قلب', 'إنعاش',
                                                   'انعاش', 'القلب', 'قلبي']):
                            timer_needed = dict(duration=120, title='❤️ CPR Cycle Timer',
                                                instructions='30 compressions, then 2 breaths - repeat')
                        elif any(w in ulq for w in ['burn', 'حرق', 'حروق', 'الحروق',
                                                     'brûlure', 'quemadura']):
                            timer_needed = dict(duration=1200, title='🔥 Burn Cooling Timer',
                                                instructions='Keep burn under cool running water')
                        elif any(w in ulq for w in ['chok', 'اختناق', 'الاختناق', 'étouff', 'asfixia']):
                            timer_needed = dict(duration=300, title='🫁 Choking Response Timer',
                                                instructions='Continue attempts until object dislodges')
    
                if timer_needed:
                    # If the chatbot returned a low-confidence message but the
                    # timer fired, a match WAS found — replace the vague reply
                    # with the actual first-aid answer before appending the timer.
                    _LOW_CONF_MARKERS = [
                        "I found something related, but I'm not confident",
                        "I'm not entirely sure what situation you're describing",
                        "I'm sorry, I couldn't find a good match",
                    ]
                    if any(marker in bot_response for marker in _LOW_CONF_MARKERS):
                        _underlying = getattr(st.session_state.chatbot, 'chatbot', None)
                        if _underlying is not None:
                            _em, _conf = _underlying.find_best_match(
                                user_input, threshold=0.0
                            )
                            if _em is not None:
                                bot_response = _underlying.format_answer_smart(
                                    _em, user_input, _conf
                                )
                                print(
                                    f"[Timer override] score={_conf:.4f} — replaced "
                                    f"low-confidence response with full answer for: "
                                    f"{user_input!r}"
                                )

                    st.session_state.active_timer = timer_needed
                    st.session_state.timer_start_time = datetime.datetime.now()
                    st.session_state.timer_paused = True
                    st.session_state.timer_paused_remaining = timer_needed['duration']
                    user_lang = getattr(st.session_state.chatbot, 'user_language', 'en')
                    timer_msg = "\n" + get_timer_instruction_translated(
                        timer_needed['title'], user_lang)
                    bot_response += timer_msg
                    st.session_state.chat_history[-1]['content'] = bot_response
                    if st.session_state.logged_in:
                        auth.save_user_chat(st.session_state.username, st.session_state.chat_history)
    
            st.rerun()
    

# ─── Router ───────────────────────────────────────────────────────────────────
def main():
    if st.session_state.page == 'landing':
        show_landing_page()
    elif st.session_state.page == 'login':
        show_login_page()
    elif st.session_state.page == 'signup':
        show_signup_page()
    elif st.session_state.page == 'chat':
        show_chat_page()


if __name__ == "__main__":
    main()
